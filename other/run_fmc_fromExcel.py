# Run final mass calculation using collated data from an Excel file

import os
import numpy as np
from openpyxl.styles import Font, Alignment

from msl.io import read_table

from mass_circular_weighing.configuration import Configuration
from mass_circular_weighing.constants import MU_STR

from mass_circular_weighing.routine_classes.final_mass_calc_class import FinalMassCalc
from mass_circular_weighing.gui.threads.masscalc_popup import filter_mass_set

from mass_circular_weighing.routine_classes.results_summary_Excel import ExcelSummaryWorkbook


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
admin = r'C:\Users\r.hawke\OneDrive - Callaghan Innovation\Desktop\Recal_RJH\MassStdsD10_Admin.xlsx'
# input_data is in admin file

# test data
data_table = read_table(admin, sheet='InputData')

collated = np.empty(len(data_table),
                    dtype=[('+ weight group', object), ('- weight group', object),
                           ('mass difference (g)', 'float64'), ('balance uncertainty (' + MU_STR + 'g)', 'float64'),
                           ('air density (kg/m3)', 'float64'), ('uncertainty (kg/m3)', 'float64'),
                           ],
                    )
collated['+ weight group'] = data_table[:, 0]
collated['- weight group'] = data_table[:, 1]
for i in range(len(data_table)):
    collated['mass difference (g)'][i] = float(data_table[i, 2])
    collated['balance uncertainty (' + MU_STR + 'g)'][i] = float(data_table[i, 3])
    collated['air density (kg/m3)'][i] = float(data_table[i, 4])
    collated['uncertainty (kg/m3)'][i] = float(data_table[i, 5])


cfg = Configuration(admin)
cfg.init_ref_mass_sets()

print(cfg.all_checks)
print(cfg.all_stds)

client_wts = filter_mass_set(cfg.all_client_wts, collated)
checks = None  # filter_mass_set(cfg.all_checks, collated)
stds = filter_mass_set(cfg.all_stds, collated)

fmc = FinalMassCalc(cfg.folder, cfg.client, client_wts, checks, stds, collated, nbc=False, corr=cfg.correlations)


# settings for analysis options
fmc.BUOYANCY_CORR = True
fmc.HEIGHT_CORR = False

fmc.UNC_AIR_DENS = True
fmc.UNC_VOL = False
fmc.UNC_HEIGHT = False

fmc.TRUE_MASS = False


# do calculation
fmc.import_mass_lists()
fmc.parse_inputdata_to_matrices()

# fmc.apply_corrections_to_mass_differences(air_densities=collated['air density (kg/m3)'])
# fmc.cal_psi_y(unc_airdens=collated['uncertainty (kg/m3)'], air_densities=air_densities, rv=None)

fmc.do_least_squares()
fmc.check_residuals()

fmc.cal_rel_unc()
fmc.make_summary_table()

covariance = fmc.covariances[0, 1]
print(covariance)
variance = np.sqrt(fmc.covariances[0, 0] * fmc.covariances[1, 1])
print(variance)
corr_coeff = covariance/variance
print(corr_coeff)

fmc.add_data_to_root()
fmc.save_to_json_file()

# Export to Excel summary file
xl = ExcelSummaryWorkbook(cfg)
xl.format_scheme_file()
xl.add_mls(fmc.finalmasscalc)

# add air density data to inputdata sheet
insheet = xl.wb["MLS Input Data"]
insheet["F3"] = "Air Density (kg/m3)"
for i, value in enumerate(data_table[:, 4]):
    insheet.cell(row=i + 4, column=6).value = float(value)
    insheet.cell(row=i + 4, column=6).number_format = "0.000 000"
insheet["G3"] = "uncertainty (kg/m3)"
for i, value in enumerate(data_table[:, 5]):
    insheet.cell(row=i + 4, column=7).value = float(value)
for cell in [insheet["F3"], insheet["G3"]]:
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                               shrink_to_fit=False, indent=0)


# add covariance matrix info to mls sheet
mls = xl.wb["MLS Output Data"]
mls.append([])  # Makes a new empty row
mls.append(["Covariance matrix in g"])
mls.append([" "]+fmc.all_wts['Weight ID'])
for i, row in enumerate(fmc.covariances):
    mls.append([fmc.all_wts['Weight ID'][i]]+list(row))

xl_output_file = xl.save_workbook(cfg.folder, cfg.client)
# set Excel output to read_only
# os.chmod(xl_output_file, S_IREAD | S_IRGRP | S_IROTH)
