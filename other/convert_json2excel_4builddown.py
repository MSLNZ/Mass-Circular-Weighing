"""
A helper script to convert a json file from a set of circular weighings
into a table of weighing data for use in the builddown analysis.
"""
import os
import string
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from msl.io import read
from msl.equipment import Config

from mass_circular_weighing.constants import IN_DEGREES_C, DEGREE_SIGN
from mass_circular_weighing.log import log


folder = r'C:\Users\rebecca.hawke\Desktop\builddown testing'  # folder of data
json_file = r'C:\Users\rebecca.hawke\Desktop\builddown testing\MassStds_1000.json'

jf = read(json_file)
print(jf)

se = "1KSR 1KSS 1KSC 1KCC"
cw_file = json_file
nom = '1000'
cfg = Config(r"C:\MCW_Config\local_config.xml")


def add_weighing_dataset(se, cw_file, nom, cfg):
    """Adds relevant from each circular weighing for a given scheme entry

        Parameters
        ----------
        se : str
        cw_file : path
        nom : str
        incl_datasets : set
        cfg : configuration instance
    """
    wb = Workbook()     # create new Excel doc
    sheet = wb.active   # get active sheet
    if not os.path.isfile(cw_file):
        log.warning(f'No data yet collected for {se}')
    else:
        wt_grps = se.split()

        root = read(cw_file)
        log.info(f"Adding data for {se} from {cw_file}")
        sheet.append([f'Circular weighings for {se}'])
        sheet['A1'].font = Font(bold=True)
        sheet.append([f'Source file: {cw_file}'])

        try:
            root['Circular Weighings'][se]
        except KeyError:
            log.warning(f'No data yet collected for {se}')
            return

        first_dataset = True
        for dataset in root['Circular Weighings'][se].datasets():
            dname = dataset.name.split('_')

            if dname[0][-8:] == 'analysis':
                run_id = 'run_' + dname[2]
                run_no = float(run_id.strip('run_'))

                weighdata = root.require_dataset(
                    root['Circular Weighings'][se].name + '/measurement_' + run_id)

                try:
                    # TODO calculate actual mean and save to metadata as part of run_circ_weigh
                    temps = weighdata.metadata.get("T" + IN_DEGREES_C).split(" to ")
                    mean_temp = (float(temps[0]) + float(temps[1])) / 2
                    rhs = weighdata.metadata.get("RH (%)").split(" to ")
                    mean_rhs = (float(rhs[0]) + float(rhs[1])) / 2
                except AttributeError:
                    temps = []
                    mean_temp = []
                    rhs = []
                    mean_rhs = []

                incl = 1

                # Get balance model number from balance alias:
                bal_alias = weighdata.metadata.get("Balance")
                db = cfg.database()
                bal_model = db.equipment[bal_alias].model
                analysisdata = root.require_dataset(
                    root['Circular Weighings'][se].name + '/analysis_' + run_id)

                if first_dataset:
                    # get information for header lines
                    sheet.append([])
                    sheet.append(['Balance', bal_model, 'Unit', weighdata.metadata.get("Unit")])
                    sheet.append([])
                    sheet.append(["Weight group loading order"])
                    pos = []
                    for key, value in weighdata.metadata.get("Weight group loading order").items():
                        sheet.append([key, value])
                        pos.append(key.strip("Position "))

                    header_pos = ["", "", ""]
                    for p in range(len(wt_grps) - 1):
                        key = pos[p] + ' - ' + pos[p + 1]
                        header_pos.append(key)
                    header_pos.append(pos[-1] + ' - ' + pos[0])
                    sheet.append(header_pos)

                    header = [
                        "Timestamp",
                        'Run #',
                        "Included?"
                    ]
                    for row in analysisdata:  # get pairwise comparison names
                        header.append(f"({row[0]}) - ({row[1]})")
                    header += [
                        'Drift type',
                        'Residual Std. Dev.',
                        'Ambient OK?',
                        "min T" + IN_DEGREES_C,
                        "max T" + IN_DEGREES_C,
                        "mean T" + IN_DEGREES_C,
                        "min RH (%)",
                        "max RH (%)",
                        "mean RH (%)",
                        "start P (hPa)",
                        "end P (hPa)",
                        "mean P (hPa)",
                        "Balance",
                        "Unit",
                    ]
                    sheet.append(header)

                    first_dataset = False

                # add data for each run
                data_list = [
                    weighdata.metadata.get("Mmt Timestamp"),
                    run_no,
                    incl
                ]
                for row in analysisdata:  # get pairwise comparison differences
                    data_list.append(row[2])

                drift = analysisdata.metadata.get("Selected drift")
                res_dict = json.loads(str(analysisdata.metadata.get("Residual std devs")).replace("'", '"'))
                res = res_dict.get(drift)

                data_list += [
                    drift,
                    res,
                    str(weighdata.metadata.get("Ambient OK?"))
                ]

                if temps:
                    data_list += [
                        min(temps),
                        max(temps),
                        mean_temp,
                        min(rhs),
                        max(rhs),
                        mean_rhs,
                    ]
                else:
                    data_list += ["", "", "", "", "", ""]

                try:
                    p = weighdata.metadata.get("Pressure (hPa)").split(" to ")
                    mean_P = (float(p[0]) + float(p[1])) / 2
                    data_list += [p[0], p[1], mean_P]
                except AttributeError:
                    data_list += ["", "", ""]
                except TypeError:
                    data_list += ["", "", ""]

                data_list += [bal_model, weighdata.metadata.get("Unit")]

                sheet.append(data_list)

        sheet['A6'].font = Font(italic=True)
        sheet.column_dimensions["A"].width = 15

        for cell in sheet[8 + len(wt_grps)]:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)

        # determine duration of weighing in hours
        duration_formula = "=(A" + str(sheet.max_row) + " - A" + str(13) + ")*24"
        # calculate dT/dt over all runs
        # Find column of mean temperatures
        mT = string.ascii_lowercase[8 + len(wt_grps)]
        dTdt_formula = "=(" + mT + str(sheet.max_row) + " - " + mT + str(13) + ")/A" + str(sheet.max_row + 2)

        # determine formulae for mean and std dev of included weighings
        ave_row = [duration_formula, "hrs", "Mean"]
        stdev_row = ["", "", "Std. Dev."]
        for col in string.ascii_lowercase[3:3 + len(wt_grps)]:  # columns of weighing differences
            formula_ave = "=AVERAGE("
            formula_stdev = "=STDEV("
            for i, cell in enumerate(sheet['C']):
                if cell.value == 1:  # included weighing
                    cell_name = col + str(i + 1) + ','  # enumerate starts from 0, rows from 1
                    formula_ave += cell_name
                    formula_stdev += cell_name
            if formula_ave == "=AVERAGE(":
                log.warning(f"No data selected for {se}")
                break
            formula_ave = formula_ave.strip(',') + ")"
            formula_stdev = formula_stdev.strip(',') + ")"
            ave_row.append(formula_ave)
            stdev_row.append(formula_stdev)

        for i in range(5):
            ave_row.append(" ")
        ave_row.append(dTdt_formula)
        ave_row.append(DEGREE_SIGN+"C/h")

        sheet.append([])
        sheet.append(ave_row)
        for cell in sheet[sheet.max_row]:
            cell.font = Font(italic=True)
        sheet.append(stdev_row)
        for cell in sheet[sheet.max_row]:
            cell.font = Font(italic=True)

        sheet.append([])

        wb.save(os.path.join(folder, "summary.xlsx"))


add_weighing_dataset(se, cw_file, nom, cfg)

#
# summary = Workbook()                            # create new Excel doc
# runs = summary.active                           # get active sheet
# runs.title = "Individual Runs"                  # set title
# runs.append([
#     "File", "Mmt Date", "Nominal (g)", "Run #", 'no drift', 'linear drift', 'quadratic drift', 'cubic drift',
# ])
# collated = summary.create_sheet("Collated")     # make new sheet for collated runs
# collated.append([
#     "File", "Mmt Date", "Nominal (g)", "Scheme entry", "Run #", "+ weight group", "- weight group",
#     "mass difference (g)", "residual (µg)", "balance uncertainty (µg)", "Acceptance met?",
# ])
