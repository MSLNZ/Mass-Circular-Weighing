"""
A script to collate all weighing data from json files to determine acceptance criteria
"""

import os
from openpyxl import Workbook
import numpy as np
from datetime import datetime

from msl.io import read

# from mass_circular_weighing.constants import MU_STR, SUFFIX

folder = r'M:\Commercial Calibrations'  #\2025\0011_Glycosyn'  #\UMX-5\Linearity'  # folder of data

summary = Workbook()                            # create new Excel doc
summary.iso_dates = True                        # store dates and times in the ISO8601 format
runs = summary.active                           # get active sheet
runs.title = "Individual Runs"                  # set title
runs.append([
    "File", "Mmt Date", "Bal", "Nominal (g)", "Scheme Entry", "Run #", "no drift", "linear drift", 'quadratic drift', 'cubic drift',
])
collated = summary.create_sheet("Collated")     # make new sheet for collated runs
collated.append([
    "File", "Mmt Date", "Bal", "Nominal (g)", "Scheme entry", "Run #", "+ weight group", "- weight group",
    "mass difference (g)", "residual (µg)", "balance uncertainty (µg)", "Acceptance met?",
])

for flist in os.walk(folder):                   # traverse all folders and subfolders
    if 'executable_archives' in flist[0]:
        print('executables')
    elif 'ackup' not in flist[0]:  # unless they are called 'backups' or 'Backups'
        for fname in flist[2]:                  # iterate through the files
            if '.json' in fname and 'finalmasscalc' not in fname:   # select only json files from weighings
                file_to_read = os.path.join(flist[0], fname)
                print(f'Reading {file_to_read}')
                jsonroot = read(file_to_read)
                nom = None
                date = None
                for ds in jsonroot.datasets():
                    if 'measurement_run_1' in ds.name:
                        nom = ds.metadata['Nominal mass (g)']
                if 110 < nom < 550:
                    # print(nom, file_to_read.strip(folder))
                    for ds in jsonroot.datasets():
                        if 'measurement_run_1' in ds.name:
                            bal = ds.metadata["Balance"]
                        if 'analysis' in ds.name:
                            mmt = ds.name.replace('analysis', 'measurement')
                            mmt_ds = jsonroot[mmt]
                            date = mmt_ds.metadata["Mmt Timestamp"]
                            run_data = [
                                file_to_read.strip(folder),
                                date,
                                mmt_ds.metadata["Balance"],
                                nom,
                                ds.name.split("/")[2],  # scheme entry
                                ds.name.split("_")[-1]
                            ]
                            res_dict = eval(ds.metadata["Residual std devs"])
                            for key, value in res_dict.items():
                                run_data.append(value)
                            # print(run_data)
                            runs.append(run_data)
                        if 'Collated' in ds.name:
                            mmt = ds.name.replace('Collated', 'measurement_run_1')
                            mmt_ds = jsonroot[mmt]
                            date = mmt_ds.metadata["Mmt Timestamp"]
                            bal = mmt_ds.metadata["Balance"]
                            for r in ds.data:
                                row = [file_to_read.strip(folder), date, bal]
                                for i in r:
                                    row.append(i)
                                collated.append(row)

# add column, and split date and time
runs.insert_cols(3, amount=1)
runs['C1'] = "Mmt Time"
for i in range(2, runs.max_row+1):
    try:
        date_time = datetime.strptime(runs['B'+str(i)].value, '%d-%m-%Y %H:%M:%S')
    except ValueError:
        date_time = datetime.strptime(runs['B' + str(i)].value, '%d-%m-%Y %H:%M')
    runs['B' + str(i)] = date_time.date()
    runs['C' + str(i)] = date_time.time()

collated.insert_cols(3, amount=1)
collated['C1'] = "Mmt Time"
for i in range(2, collated.max_row+1):
    date_time = collated['B'+str(i)].value.split()
    collated['B' + str(i)] = date_time[0].replace("-", "/")
    collated['C' + str(i)] = date_time[1]

# save data to file
summary.save(os.path.join(folder, "summary.xlsx"))
