"""
Results summary in Excel format
Called from routines/report_results.py
"""
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from msl.io import read

from .constants import IN_DEGREES_C
from .log import log


def list_to_csstr(idlst):
    idstr = ""
    for id in idlst:
        idstr += id + ", "
    return idstr.strip(" ").strip(",")


class ExcelSummaryWorkbook(object):

    def __init__(self):
        self.wb = None
        self.first_scheme_entry_row = 2
        self.collate_ambient = {'T' + IN_DEGREES_C: [], 'RH (%)': []}

    def load_scheme_file(self, scheme_file_name, fmc_root, check_file, std_file, job, client, folder):
        "Begins the summary file with the weighing scheme, as saved from the main gui"
        wb = load_workbook(scheme_file_name)
        self.wb = wb
        scheme_sheet = self.wb["Scheme"]
        for cell in scheme_sheet[1]:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)
        last_row = scheme_sheet.max_row

        scheme_sheet.append(["Summary of weighings"])
        scheme_sheet.append(["Job", job])
        scheme_sheet.append(["Client", client])
        scheme_sheet.append(["Folder of weighing data", folder])
        scheme_sheet.append([])
        scheme_sheet.append(["Weight sets"])

        client_wt_IDs = list_to_csstr(fmc_root["1: Mass Sets"]["Client"].metadata.get("weight ID"))
        scheme_sheet.append(['Client weights', client_wt_IDs])

        num_new_rows = 12

        if check_file:
            check_wt_IDs = list_to_csstr(fmc_root["1: Mass Sets"]["Check"].metadata.get("weight ID"))
            scheme_sheet.append(['Check weights', check_wt_IDs])
            scheme_sheet.append(["", check_file])
            num_new_rows += 1
        else:
            scheme_sheet.append(['Check weights', "None"])

        std_wts = list_to_csstr(fmc_root["1: Mass Sets"]["Standard"].metadata.get("weight ID"))
        scheme_sheet.append(['Standard weights', std_wts])
        scheme_sheet.append(["", std_file])
        scheme_sheet.append([])
        scheme_sheet.append(["Weighing scheme"])

        scheme_sheet.insert_rows(0, num_new_rows)
        scheme_sheet.move_range("A"+str(last_row+num_new_rows+1)+":C"+str(last_row+num_new_rows*2), -(last_row+num_new_rows))
        self.first_scheme_entry_row = num_new_rows + 2

        scheme_sheet['A1'].font = Font(bold=True)
        scheme_sheet['A6'].font = Font(bold=True)
        scheme_sheet['A'+str(num_new_rows)].font = Font(bold=True)
        scheme_sheet.column_dimensions["A"].width = 21

    def save_array_to_sheet(self, data, sheet_name):
        "Quick method to dump a NumPy array into an Excel sheet. Requires metadata of array to be column headers"
        sheet = self.wb.create_sheet(sheet_name)

        header = data.metadata.get('metadata')['headers']
        sheet.append(header)

        for row in data:
            sheet.append(list(row))

        for cell in sheet[1]:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)

    def add_mls(self, fmc_root):
        """Adds matrix least squares sections to summary file;
        separates input and output data into two different sheets"""

        # Save output to sheet
        outdata = fmc_root['2: Matrix Least Squares Analysis']["Mass values from least squares solution"]
        self.save_array_to_sheet(outdata, sheet_name="MLS Output Data")

        mls = self.wb["MLS Output Data"]
        mls.insert_rows(0, 3)
        mls['A1'] = 'Mass values from Least Squares solution'
        mls['A2'] = fmc_root['metadata'].metadata['Timestamp']
        mls['A1'].font = Font(bold=True)

        cols = ["A", "B", "D", "E", "H"]
        widths = [16, 11, 16, 16, 30]
        for col, width in zip(cols, widths):
            mls.column_dimensions[col].width = width

        # add metadata
        meta = fmc_root['2: Matrix Least Squares Analysis']['metadata'].metadata
        mls.append([])  # Makes a new empty row
        for key, value in meta.items():
            mls.append([key, value])
            cell = mls["A"+str(mls.max_row)]
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)

        # add custom number formatting
        for i in range(5, mls.max_row):
            mls["D"+str(i)].number_format = "0.000 000 000"

        # Save input to sheet
        indata = fmc_root['2: Matrix Least Squares Analysis']["Input data with least squares residuals"]
        self.save_array_to_sheet(indata, sheet_name="MLS Input Data")
        insheet = self.wb["MLS Input Data"]
        insheet.insert_rows(0, 2)
        insheet['A1'] = "Input data for Matrix Least Squares analysis"
        insheet['A1'].font = Font(bold=True)

        # add custom number formatting
        for i in range(4, insheet.max_row):
            insheet["C"+str(i)].number_format = "0.000 000 000"
        insheet.column_dimensions["A"].width = 15
        insheet.column_dimensions["B"].width = 15
        insheet.column_dimensions["C"].width = 16
        insheet.column_dimensions["D"].width = 19

    def add_weighing_dataset(self, se, cw_file,  nom, incl_datasets, cfg):
        """Adds relevant from each circular weighing for a given scheme entry

                Parameters
                ----------
                se : str
                cw_file : path
                nom : str
                incl_datasets : set
                cfg : configuration instance
                """
        if not os.path.isfile(cw_file):
            log.warning('No data yet collected for ' + se)
        else:
            if len(se) > 30:
                # MS Excel has a limit of 31 characters in tab name
                sheet = self.wb.create_sheet(se[:27]+"...")
            else:
                sheet = self.wb.create_sheet(se)
            wt_grps = se.split()

            root = read(cw_file)
            log.info("Adding data for " + se + " from " + cw_file)
            sheet.append(["Circular weighings for "+se])
            sheet['A1'].font = Font(bold=True)
            sheet.append(['Source file: '+cw_file])

            try:
                root['Circular Weighings'][se]
            except KeyError:
                log.warning('No data yet collected for '+se)
                return

            for dataset in root['Circular Weighings'][se].datasets():
                dname = dataset.name.split('_')

                if dname[0][-8:] == 'analysis':
                    run_id = 'run_' + dname[2]
                    run_no = float(run_id.strip('run_'))

                    weighdata = root.require_dataset(
                        root['Circular Weighings'][se].name + '/measurement_' + run_id)

                    temps = weighdata.metadata.get("T" + IN_DEGREES_C).split(" to ")
                    rhs = weighdata.metadata.get("RH (%)").split(" to ")

                    if (str(float(nom)), se, dname[2]) in incl_datasets:
                        incl = 1
                        for t in temps:
                            self.collate_ambient['T' + IN_DEGREES_C].append(float(t))
                        for rh in rhs:
                            self.collate_ambient['RH (%)'].append(float(rh))
                    else:
                        incl = 0

                    # Get balance model number from balance alias:
                    bal_alias = weighdata.metadata.get("Balance")
                    bal_model = cfg.equipment[bal_alias].model
                    analysisdata = root.require_dataset(
                        root['Circular Weighings'][se].name + '/analysis_' + run_id)

                    if run_no < 2:
                        # get information for header lines
                        sheet.append([])
                        sheet.append(['Balance', bal_model, 'Unit', weighdata.metadata.get("Unit")])
                        sheet.append([])
                        sheet.append(["Weight group loading order"])
                        pos = []
                        for key, value in weighdata.metadata.get("Weight group loading order").items():
                            sheet.append([key, value])
                            pos.append(key.strip("Position "))

                        header_pos = ["", "", ""]
                        for p in range(len(wt_grps) - 1):
                            key = pos[p] + ' - ' + pos[p+1]
                            header_pos.append(key)
                        header_pos.append(pos[-1] + ' - ' + pos[0])
                        sheet.append(header_pos)

                        header = [
                            "Timestamp",
                            'Run #',
                            "Included?"
                        ]
                        for row in analysisdata:
                            header.append("("+row[0] + ") - (" + row[1]+")")
                        header += [
                            'Drift type',
                            'Residual Std. Dev.',
                            'Ambient OK?',
                            "min T" + IN_DEGREES_C,
                            "max T" + IN_DEGREES_C,
                            "min RH (%)",
                            "max RH (%)",
                            "start P (hPa)",
                            "end P (hPa)",
                        ]
                        sheet.append(header)

                    # add data for each run
                    data_list = [
                        weighdata.metadata.get("Mmt Timestamp"),
                        run_no,
                        incl
                    ]
                    for row in analysisdata:
                        data_list.append(row[2])

                    drift = analysisdata.metadata.get("Selected drift")
                    res_dict = json.loads(str(analysisdata.metadata.get("Residual std devs")).replace("'", '"'))
                    res = res_dict.get(drift)

                    data_list += [
                        drift,
                        res,
                        str(weighdata.metadata.get("Ambient OK?")),
                        min(temps),
                        max(temps),
                        min(rhs),
                        max(rhs),
                    ]

                    try:
                        # add pressures if recorded - need to check this!
                        p = weighdata.metadata.get("Pressure (hPa)")
                        data_list += [p[0], p[1]]
                    except TypeError:
                        pass

                    sheet.append(data_list)

            sheet['A6'].font = Font(italic=True)
            sheet.column_dimensions["A"].width = 15

            for cell in sheet[8+len(wt_grps)]:
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                        shrink_to_fit=False, indent=0)

    def add_all_cwdata(self, cfg, incl_datasets,):
        scheme = self.wb['Scheme']
        i = self.first_scheme_entry_row  # job/client etc info has been added before scheme

        while True:
            se = scheme.cell(row=i, column=1).value
            nom = scheme.cell(row=i, column=2).value
            try:
                cw_file = os.path.join(cfg.folder, cfg.client + '_' + nom + '.json')
            except TypeError:
                break
            self.add_weighing_dataset(se, cw_file, nom, incl_datasets, cfg)
            i += 1

    def add_overall_ambient(self):
        mls = self.wb["MLS Output Data"]
        mls.append([])
        mls.append(["Overall range of ambient conditions for included datasets"])
        mls['A'+str(mls.max_row)].font = Font(italic=True)
        for key, value in self.collate_ambient.items():
            try:
                mls.append([key, min(value), max(value)])
            except ValueError:
                mls.append([key, "(no data)"])

    def save_workbook(self, path):
        self.wb.save(path)
        log.info('Data saved to {}'.format(path))
