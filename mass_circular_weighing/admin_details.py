"""
This class requires an .xlsx file in the correct template
"""
import os
import io
import string
import numpy as np

from openpyxl import load_workbook, Workbook

from .log import log
from .constants import config_default, client_default, job_default, MU_STR
from .gui.threads.prompt_thread import PromptThread
pt = PromptThread()

header_row = 14


class AdminDetails(object):

    def __init__(self, path):
        """Administrative details for a calibration are read from an .xlsx file.
        Essential details are checked, and default values used if needed.
        Details of client masses and standard mass sets are stored in dictionary objects.

        Parameters
        ----------
        path : path
            an .xlsx file in the correct template. Use examples/Admin.xlsx as the template.
        """
        self.path = path

        with open(path, "rb") as f:         # so that the file remains available to be edited after being read
            self.wb = load_workbook(f, data_only=True)
        log.info(f"Found Admin file at {self.path}")

        self.ds = self.wb["Admin"]  # note that this will raise an error if the sheet Admin doesn't exist

        self.operator = self.ds['B2'].value

        # Client
        try:
            # remove any spaces if present
            self.client = self.ds['B3'].value.replace(" ", "")
        except AttributeError:
            # no spaces present to remove
            self.client = self.ds['B3'].value
        if not self.client:
            self.client = client_default
            self.ds['B3'] = self.client  # update value in memory
            log.warning(f"No client name specified. Defaulting to {self.client}.")

        # Job
        self.job = self.ds['B4'].value
        if not self.job:
            self.job = job_default
            self.ds['B4'] = self.job
            log.warning(f"No job number specified. Defaulting to {self.job}.")

        # Save Folder
        try:
            self.folder = self.ds['B5'].value.encode('unicode-escape').decode()  # convert to raw string
            if not os.path.exists(self.folder):
                os.makedirs(self.folder)
        except AttributeError:  # no folder specified
            # get folder information from where the Admin.xlsx file is coming from
            self.folder = os.path.dirname(self.path)  # save_folder_default
            self.ds['B5'] = self.folder
            log.warning(f"No save folder specified. Defaulting to {self.folder}.")

        # Configuration File
        self.config_xml = self.ds['E11'].value
        if not self.config_xml:
            # look for a config file in the same folder as the Admin.xlsx file
            xml_files = [f for f in os.listdir(self.folder) if f.endswith(".xml")]
            if xml_files:
                pt = PromptThread()
                pt.show('item', "Select your config.xml file if present", xml_files)
                config_xml = pt.wait_for_prompt_reply()
                if config_xml:
                    self.config_xml = os.path.join(self.folder, config_xml)
                    log.warning(f"No config.xml file path specified in {self.path}.")
            if not self.config_xml:
                self.config_xml = config_default  # the example config file in Mass-Circular-Weighing
                log.warning(f"No config.xml file path specified. Defaulting to {self.config_xml}.")
        if not os.path.isfile(self.config_xml):
            raise FileNotFoundError(f"Cannot find the configuration file at {self.config_xml}.")
        self.ds['E11'] = self.config_xml

        # Circular Weighing Analysis Parameters
        self.drift_text = self.ds['E7'].value
        self.timed_text = self.ds['E8'].value
        self.calc_true_mass = self.ds['E9'].value

        # correlations are included as a 2x2 matrix - if no values are found, the identity matrix is used
        try:
            self.correlations = np.array([[float(i.value) for i in j] for j in self.ds['I8:J9']])
            print(self.correlations)
            log.info(f'Using matrix of correlations:\n{self.correlations}')
        except TypeError:
            self.correlations = np.identity(2)
            log.info(f'No correlations between standards')

        # Weight Set Information
        self.all_client_wts = self.load_client_set()
        self.client_wt_IDs = self.all_client_wts['Weight ID']

        self.all_stds = None
        self.all_checks = None
        self.massref_path = self.ds['B9'].value
        if not os.path.isfile(self.massref_path):
            # open a browser to find the MASSREF file
            pt = PromptThread()
            pt.show('filename', title="Please select a valid MASSREF file", filters='XLSX files (*.xlsx)', multiple=False)
            self.massref_path = pt.wait_for_prompt_reply()

            if not os.path.isfile(self.massref_path):
                raise FileNotFoundError(f"Cannot find the MASSREF file at {self.massref_path}.")
        log.info(f"Found MassRef file at {self.massref_path}")

        self.std_set = self.ds['B10'].value
        if not self.std_set:
            log.error("No reference mass set specified!")
        self.check_set_text = self.ds['B11'].value

        self.scheme = self.load_scheme()

    @property
    def drift(self) -> str | None:
        """Allowed options for returned string are: 'no drift', 'linear drift', 'quadratic drift', 'cubic drift'.
        If 'auto select', drift is set to `None` such that the analysis routine selects optimal drift correction.
        """
        if self.drift_text == 'auto select':
            return None
        return self.drift_text

    @property
    def timed(self) -> bool:
        """if `True`, analysis uses times from weighings, otherwise assumes equally spaced in time"""
        if self.timed_text == 'NO':
            return False
        return True

    @property
    def check_set(self):
        if self.check_set_text == 'None':
            return None
        return self.check_set_text

    def load_scheme(self):
        """Loads the weighing scheme from the 'Scheme' sheet of an .xlsx file if present"""
        try:
            sheet = self.wb["Scheme"]
        except KeyError:
            log.info('Scheme worksheet does not yet exist in {}'.format(self.path))
            return None

        header = []
        rows = []
        for i, row in enumerate(sheet.values):
            if i == 0:
                header = [item for item in row]
            else:
                rows.append([item for item in row])

        return header, rows

    def load_client_set(self):
        """Collects relevant weight info for all weights in set

        Returns
        -------
        dict
            keys:
                'Set identifier', 'Client', 'Weight ID', 'Nominal (g)', 'Shape/Mark', 'Container',
                'u_mag (mg)', 'Density (kg/m3)', 'u_density (kg/m3)', 'Expansion coeff (ppm/degC)',
                'Num weights', 'Vol (mL)'
        """
        wt_dict = {
            'Set identifier': None,
            'Set type': 'Client',
            'Client': self.client,
            'Weight ID': []
        }

        col_name_keys = {
            "weight id": 'Weight ID', "nom": 'Nominal (g)', 'mark': 'Shape/Mark', "container": 'Container',
            'u_mag': 'u_mag (mg)', 'density': 'Density (kg/m3)', 'u_dens': 'u_density (kg/m3)',
            'expans': 'Expansion coeff (ppm/degC)', 'centre height': 'Centre Height (mm)', 'u_height': 'u_height (mm)'
        }   # warning: 'u_density' contains 'density' so always use 'u_dens' in xlsx file instead.

        for i in string.ascii_uppercase[:10]:  # go across a row ;)
            key = self.ds[i+str(header_row)].value
            # do a look up to make sure the column name is a valid key, and use the valid key instead
            valid = False
            for code, real_key in col_name_keys.items():
                if code in key.lower():
                    key = real_key
                    valid = True
            if not valid:
                raise ValueError(f'Error in parsing client weight set: {key} not recognised as a known column header')
            # key is valid so we can parse the data in that column
            val = []
            for row in range(header_row + 1, self.ds.max_row):
                v = self.ds[i + str(row)].value
                if key == 'Weight ID':
                    if v is None:
                        break
                    else:
                        # ensure all weight IDs are strings
                        val.append(str(v))
                else:
                    # keep whatever data type makes sense
                    val.append(v)

                wt_dict[key] = val

        if not wt_dict['Weight ID']:
            log.error("No weights in client weight set!")
        # make sure all the lists are the same length as the Weight ID list
        num_weights = len(wt_dict['Weight ID'])
        for key, val in col_name_keys.items():
            wt_dict[val] = wt_dict[val][:num_weights]
        wt_dict['Num weights'] = num_weights

        # calculate volumes if we have the densities
        add_volumes(wt_dict)

        return wt_dict

    def init_ref_mass_sets(self) -> None:
        """Collects relevant weight info for all reference and check sets following ':meth:`.load_set_from_massref
        """
        with open(self.massref_path, "rb") as f:  # so that the file remains available to be edited after being read
            massrefwb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)

            self.all_stds = self.load_set_from_massref(massrefwb, sheet=self.std_set, set_ID="Standard")

            if self.check_set is not None:
                self.all_checks = self.load_set_from_massref(massrefwb, sheet=self.check_set, set_ID='Check')
            else:
                self.all_checks = None

    def load_set_from_massref(self, massrefwb: Workbook, sheet: str, set_ID: str) -> dict:
        """Makes a dictionary of the relevant weight info from a MASSREF.xlsx file for all weights in set

        :param massrefwb: The mass set data
        :param sheet: The sheet name for the specific mass set
        :param set_ID: The MCW program name for the mass set (client/reference/check)
        :return: A dictionary with the following keys:
            'MASSREF file', 'Sheet name', 'Set name', 'Set type', 'Set identifier', 'Calibrated',
            'Shape/Mark', 'Nominal (g)', 'Weight ID', 'mass values (g)', 'u_cal', 'uncertainties (' + MU_STR + 'g)',
            'u_drift', 'Density (kg/m3)', 'u_density (kg/m3)', 'Expansion coeff (ppm/degC)', 'Num weights', 'Vol (mL)',
            'centre height': 'Centre Height (mm)', 'u_height': 'u_height (mm)'
        """
        std_sheet = massrefwb[sheet]
        all_stds = {'MASSREF file': self.massref_path, 'Sheet name': sheet, "Set type": set_ID,
                    'Set name': std_sheet['B1'].value, 'Set identifier': std_sheet['D1'].value.strip(),
                    'Calibrated': str(std_sheet['F1'].value)}

        # use parsing of nominal values to determine last non-empty row
        for key in [
            'Shape/Mark', 'Nominal (g)', 'Weight ID', 'mass values (g)', 'u_cal', 'uncertainties (' + MU_STR + 'g)',
            'u_drift', 'Density (kg/m3)', 'u_density (kg/m3)', 'Expansion coeff (ppm/degC)',
            'Centre Height (mm)', 'u_height (mm)'
        ]:
            all_stds[key] = []

        start_row = 4
        i = 0
        while True:
            nom = std_sheet[f'B{start_row+i}'].value
            if nom is None:
                break
            if type(nom) is str:
                if nom[-1].lower() == 'k':
                    all_stds['Nominal (g)'].append(int(nom[:-1])*1000)
                else:
                    try:
                        all_stds['Nominal (g)'].append(float(nom))
                    except ValueError:
                        log.warning(f"Nominal mass {nom} not included in mass set.")
            else:
                all_stds['Nominal (g)'].append(nom)
            all_stds['Shape/Mark'].append(std_sheet[f'A{start_row + i}'].value)

            # Create weight IDs from nominal, any identifiers like d, and the set identifier.
            # Note here we are forcing k --> K
            try:
                wt_id = str(nom).upper() + std_sheet[f'C{start_row + i}'].value + all_stds['Set identifier']
            except TypeError:
                wt_id = str(nom).upper()  + all_stds['Set identifier']

            all_stds['Weight ID'].append(wt_id)
            # mass value
            mv = std_sheet[f'D{start_row + i}'].value
            all_stds['mass values (g)'].append(mv)
            # all values for uncertainty should be in micrograms
            u_cal = float(std_sheet[f'E{start_row + i}'].value)
            u_drift = float(std_sheet[f'N{start_row + i}'].value)
            u_tot = round((u_cal**2 + u_drift**2)**0.5, 3)
            all_stds['u_cal'].append(u_cal)
            all_stds['u_drift'].append(u_drift)
            all_stds['uncertainties (' + MU_STR + 'g)'].append(u_tot)
            # density and its uncertainty
            dens = std_sheet[f'G{start_row + i}'].value
            u_dens =std_sheet[f'H{start_row + i}'].value
            expans = std_sheet[f'I{start_row + i}'].value
            all_stds['Density (kg/m3)'].append(dens)
            try:    # allow u_density to be missing
                all_stds['u_density (kg/m3)'].append(float(u_dens))
            except TypeError:
                all_stds['u_density (kg/m3)'].append(None)
            try:    # allow expansion coeff to be missing
                all_stds['Expansion coeff (ppm/degC)'].append(float(expans))
            except TypeError:
                log.warning(f"No Expansion coefficient data found in the {set_ID} mass set. "
                            f"Please ensure u_drift is in column N of the appropriate MASSREF file.")
                all_stds['Expansion coeff (ppm/degC)'].append(None)

            # allow height information to be missing
            try:
                all_stds['Centre Height (mm)'].append(float(std_sheet[f'K{start_row + i}'].value))
            except TypeError:
                log.warning(f"No Centre Height data found in the {set_ID} mass set. "
                            f"Please check column K of the appropriate MASSREF file is 'Centre Height (mm)'.")
                all_stds['Centre Height (mm)'].append(None)
            try:
                all_stds['u_height (mm)'].append(float(std_sheet[f'L{start_row + i}'].value))
            except TypeError:
                log.warning(f"No Centre Height uncertainty data found in the {set_ID} mass set. "
                            f"Please check column L of the appropriate MASSREF file has the heading 'u_height (mm)'.")
                all_stds['u_height (mm)'].append(None)

            i += 1

        if i < 1:
            log.error("No weights in standard weight set!")

        all_stds['Num weights'] = len(all_stds['Weight ID'])

        # calculate volumes if we have the densities
        add_volumes(all_stds)

        return all_stds

    def save_admin(self):
        # triggered by 'Confirm settings' button on main panel of gui
        # update location of new Admin file
        self.path = os.path.join(self.folder, self.client + '_Admin.xlsx')
        # overwrite any previous version of admin details
        self.wb.save(filename=self.path)
        log.info(f'Admin details saved to {self.path}')


def add_volumes(wt_dict: dict) -> None:
    """Add a list of volumes of each weight to the weight set dictionary, using the key 'Vol (mL)'.
    If the density is provided, the volume is calculated in mL using the nominal mass of the weight.
    If the density is not provided, the volume is recorded as 'None'.

    :param wt_dict: weight set dictionary with keys as per :meth:`.load_set_from_massref` and :meth:`.load_client_set`.
    """
    vols = []
    u_vols = []
    for i in range(wt_dict['Num weights']):
        try:
            m = float(wt_dict['Nominal (g)'][i])
            d = float(wt_dict['Density (kg/m3)'][i])
            vol = 1000 * m / d
            vols.append(vol)
        except TypeError:
            vols += [None]
        try:
            m = float(wt_dict['Nominal (g)'][i])
            d = float(wt_dict['Density (kg/m3)'][i])
            u_d = float(wt_dict['u_density (kg/m3)'][i])
            rel_unc_d = u_d/d
            u_vol = rel_unc_d * 1000 * m / d
            u_vols.append(u_vol)
        except TypeError:
            u_vols += [None]

    wt_dict['Vol (mL)'] = vols
    wt_dict['Vol unc (mL)'] = u_vols

