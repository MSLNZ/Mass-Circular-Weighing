"""
A tabular display of the weighing scheme, which can import (by drag-n-drop) and export (to xlsx) Excel files.
"""
import xlrd
import openpyxl
import os
import string

from msl.qt import QtWidgets, QtCore, utils, prompt, Signal, Slot

from ...log import log


class SchemeTable(QtWidgets.QTableWidget):
    check_good_runs_in_file = Signal(int)

    def __init__(self):
        super(SchemeTable, self).__init__()
        self.bal_list = []
        self.setAcceptDrops(True)
        headers = ['Weight Groups', 'Nominal mass (g)', 'Balance alias', '# Runs', '# Collected']
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        self.make_rows(1)
        self.resizeColumnsToContents()
        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)

        verthead = self.verticalHeader()
        verthead.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        verthead.customContextMenuRequested.connect(self.vert_header_menu)

        self.scheme_path = None

    def make_rows(self, numrows):
        self.setRowCount(numrows)
        for i in range(self.rowCount()):
            self.set_cell_types(i)

    def set_cell_types(self, row_no):
        balance_io = QtWidgets.QComboBox()
        balance_io.addItems(self.bal_list)
        self.setCellWidget(row_no, 0, QtWidgets.QLineEdit())
        self.setCellWidget(row_no, 1, QtWidgets.QLineEdit())
        self.setCellWidget(row_no, 2, balance_io)
        self.setCellWidget(row_no, 3, QtWidgets.QSpinBox())
        self.setCellWidget(row_no, 4, QtWidgets.QLabel('0'))

    def update_balance_list(self, bal_list):
        """Updates list of available balances as per list in selected config.xml file, keeping current selection"""
        self.bal_list = bal_list
        for row in range(self.rowCount()):
            bal = self.cellWidget(row, 2).currentText()
            self.cellWidget(row, 2).clear()
            self.cellWidget(row, 2).addItems(self.bal_list)
            if bal in self.bal_list:
                self.cellWidget(row, 2).setCurrentIndex(self.cellWidget(row, 2).findText(bal))

    def update_status(self, row, number):
        self.cellWidget(row, 4).setText((str(number)))

    def vert_header_menu(self, pos):
        row = self.currentRow()
        if not row == -1:
            menu = QtWidgets.QMenu()
            add_above = menu.addAction("Add row above")
            add_below = menu.addAction("Add row below")
            delete = menu.addAction("Delete row")

            action = menu.exec_(self.verticalHeader().mapToGlobal(pos))

            if action == add_above:
                self.add_row_above(row)
            if action == add_below:
                self.add_row_below(row)
            if action == delete:
                self.delete_row(row)

    def add_row_above(self, row):
        self.insertRow(row)
        self.set_cell_types(row)

    def add_row_below(self, row):
        self.insertRow(row+1)
        self.set_cell_types(row+1)

    def delete_row(self, row):
        self.removeRow(row)

    def dragEnterEvent(self, event):
        paths = utils.drag_drop_paths(event, pattern='*.xls*')
        if paths:
            self.scheme_path = paths[0]
            if len(paths) > 1:
                self.scheme_path = prompt.item('Please select one file containing the weighing scheme', items=paths)
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        event.accept()

    def dropEvent(self, event):
        self.load_scheme()

    @Slot(str)
    def auto_load_scheme(self, scheme_path):
        self.scheme_path = scheme_path
        self.load_scheme()

    @Slot(list, list)
    def load_scheme(self, header=None, rows=None):
        if header is None:
            header, rows = read_excel_scheme(self.scheme_path)

        index_map = {}
        for col_name in {'weight', 'nominal', 'balance', 'runs', }:
            for i, name in enumerate(header):
                if col_name in name.lower():
                    index_map[col_name] = i

        if len(index_map) < 4:
            log.error(f"Unable to create weighing scheme from {self.scheme_path}")
            return

        self.make_rows(len(rows))
        for i, row in enumerate(rows):
            se = row[index_map['weight']]
            self.cellWidget(i, 0).setText(str(se))
            nom = row[index_map['nominal']]
            self.cellWidget(i, 1).setText(str(nom))
            self.cellWidget(i, 2).setCurrentIndex(self.cellWidget(i, 2).findText(row[index_map['balance']]))
            self.cellWidget(i, 3).setValue(int(row[index_map['runs']]))

            self.check_good_runs_in_file.emit(i)
            # updates status of number of collected runs

        if self.scheme_path:
            log.info(f'Scheme loaded from {self.scheme_path}')

            # check format is updated to xlsx
            if '.xlsx' in os.path.basename(self.scheme_path):
                return
            elif '.xls' in os.path.basename(self.scheme_path):
                self.save_scheme(os.path.dirname(self.scheme_path), os.path.basename(self.scheme_path)+'x')
                return

    def check_scheme_entries(self, cfg):
        for i in range(self.rowCount()):
            try:
                scheme_entry = self.cellWidget(i, 0).text()
                masses = []
                for wtgrp in scheme_entry.split():
                    for mass in wtgrp.split('+'):
                        masses.append(mass)
                        if mass in cfg.client_wt_IDs:
                            log.debug(mass + ' in client set')
                        elif mass in cfg.all_stds['Weight ID']:
                            log.debug(mass + ' in std set')
                        elif cfg.all_checks is not None \
                                and mass in cfg.all_checks['Weight ID']:
                            log.debug(mass + ' in check set')
                        else:
                            log.error(mass + ' is not in any of the specified mass sets')
                if len(masses) != len(set(masses)):
                    log.error(f"Duplicate masses found in {scheme_entry}")

            except AttributeError:
                pass

        log.info('Checked all scheme entries')

    def save_scheme(self, folder, filename):
        # updated to output as xlsx
        if not os.path.exists(folder):
            os.makedirs(folder)

        path = os.path.join(folder, filename)

        if os.path.isfile(path):
            workbook = openpyxl.load_workbook(path)
            try:
                # If scheme already exists, we need to clear the data
                workbook.remove(workbook["Scheme"])
            except KeyError:
                pass
            sheet = workbook.create_sheet("Scheme", 1)
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Scheme"

        header = ['Weight groups', 'Nominal mass (g)', 'Balance alias', '# runs']
        for i, a in enumerate(string.ascii_lowercase[:len(header)]):
            sheet[a+'1'] = header[i]

        for row in range(self.rowCount()):
            # scheme_entry_row = [scheme_entry, nominal, bal_alias, num_runs]
            try:
                sheet.cell(row=row+2, column=1, value=self.cellWidget(row, 0).text())
                sheet.cell(row=row+2, column=2, value=self.cellWidget(row, 1).text())
                sheet.cell(row=row+2, column=3, value=self.cellWidget(row, 2).currentText())
                sheet.cell(row=row+2, column=4, value=self.cellWidget(row, 3).text())

            except AttributeError:
                pass  #  log.error('Incomplete data in selected row')

        workbook.save(path)
        log.info('Scheme saved to ' + str(path))

    def get_se_row_dict(self, row):
        """Collate dictionary of information from selected row for weighing"""
        se_row_data = {}
        try:
            se_row_data['row'] = row
            se_row_data['scheme_entry'] = self.cellWidget(row, 0).text()
            se_row_data['nominal'] = self.cellWidget(row, 1).text()
            se_row_data['bal_alias'] = self.cellWidget(row, 2).currentText()
            se_row_data['num_runs'] = self.cellWidget(row, 3).text()

            return se_row_data

        except AttributeError:
            log.error('Incomplete data in selected row')


def read_excel_scheme(path):
    """Read an Excel file containing a weighing scheme."""
    _book = xlrd.open_workbook(path, on_demand=True)

    names = _book.sheet_names()
    if len(names) > 1:
        sheet_name = prompt.item('Please select which Sheet you wish to use:', names)
    else:
        sheet_name = names[0]

    try:
        sheet = _book.sheet_by_name(sheet_name)
    except xlrd.XLRDError:
        sheet = None

    if sheet is None:
        raise IOError('There is no Sheet named {!r} in {}'.format(sheet_name, path))

    header = [val for val in sheet.row_values(0)]
    rows = [[_cell_convert(sheet.cell(r, c)) for c in range(sheet.ncols)] for r in range(1, sheet.nrows)]
    log.debug('Loading Sheet <{}> in {!r}'.format(sheet_name, path))
    return header, rows


def _cell_convert(cell):
    """Convert an Excel cell to the appropriate value and data type"""
    t = cell.ctype
    if t == xlrd.XL_CELL_NUMBER or t == xlrd.XL_CELL_BOOLEAN:
        if int(cell.value) == cell.value:
            return '{}'.format(int(cell.value))
        else:
            return '{}'.format(cell.value)
    elif t == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code[cell.value]
    else:
        return cell.value.strip()



