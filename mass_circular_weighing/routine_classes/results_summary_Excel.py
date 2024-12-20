"""
Results summary in Excel format
Called from routines.report_results.py
"""
import os
import string
import numpy as np   # used by eval when loading residual std devs

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from msl.io import read
from ..gui.threads.prompt_thread import PromptThread
pt = PromptThread()

from ..constants import FONTSIZE, IN_DEGREES_C
from ..log import log


class ExcelSummaryWorkbook(object):

    def __init__(self, cfg):
        """Collate all administrative information, weighing data and calculated values into one spreadsheet

        Parameters
        ----------
        cfg : Config object
            Configuration class instance created using Admin.xlsx and config.xml files
        """
        # Load the Admin Details workbook which contains Admin and Scheme sheets
        self.wb = load_workbook(os.path.join(cfg.folder, cfg.client + '_Admin.xlsx'))
        if "Admin" not in self.wb.sheetnames:
            log.error(f"Admin worksheet does not exist in {cfg.path}! Please confirm settings before continuing.")
        if "Scheme" not in self.wb.sheetnames:
            log.error(f'Scheme worksheet does not exist in {cfg.path}! Please save scheme before continuing.')
        for sh in self.wb.sheetnames:
            if sh not in ["Admin", "Scheme"]:  # e.g. if the Summary file is used in place of the Admin file
                shref = self.wb.get_sheet_by_name(sh)  # get the sheet reference
                self.wb.remove_sheet(shref)  # delete the sheet

        self.first_scheme_entry_row = 2
        self.collate_ambient = {'T' + IN_DEGREES_C: [], 'RH (%)': []}

    def format_scheme_file(self):
        """Format the weighing scheme, as saved from the main gui"""
        scheme_sheet = self.wb["Scheme"]
        for cell in scheme_sheet[1]:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)

        scheme_sheet.column_dimensions["A"].width = 30

    def save_array_to_sheet(self, data, sheet_name):
        """Quick method to dump a NumPy array into an Excel sheet. Requires metadata of array to be column headers"""
        sheet = self.wb.create_sheet(sheet_name)

        header = data.metadata.get('metadata')['headers']
        sheet.append(header)

        for row in data:
            sheet.append(list(row))

        for cell in sheet[1]:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)

    def add_mls(self, fmc_root):
        """Adds matrix least squares sections to summary file;
        separates input and output data into two different sheets.
        Only seems to work with fmc_root when read from the finalmasscalc.json file using msl.io.read"""

        # Save input to sheet
        indata = fmc_root['2: Matrix Least Squares Analysis']["Input data with least squares residuals"]
        self.save_array_to_sheet(indata, sheet_name="MLS Input Data")
        insheet = self.wb["MLS Input Data"]
        insheet.insert_rows(0, 2)
        insheet['A1'] = "Input data for Matrix Least Squares analysis"
        insheet['A1'].font = Font(bold=True)

        # add custom number formatting
        for i in range(4, insheet.max_row + 1):
            insheet["C" + str(i)].number_format = "0.000 000 000"
        insheet.column_dimensions["A"].width = 15
        insheet.column_dimensions["B"].width = 15
        insheet.column_dimensions["C"].width = 16
        insheet.column_dimensions["D"].width = 19
        insheet.column_dimensions["F"].width = 19
        insheet.column_dimensions["G"].width = 19

        # Save output to sheet
        outdata = fmc_root['2: Matrix Least Squares Analysis']["Mass values from least squares solution"]
        self.save_array_to_sheet(outdata, sheet_name="MLS Output Data")

        mls = self.wb["MLS Output Data"]
        mls.insert_rows(0, 3)
        mls['A1'] = 'Mass values from Least Squares solution'
        mls['A2'] = fmc_root['metadata'].metadata['Timestamp']
        mls['A1'].font = Font(bold=True)

        cols = ["A", "B", "D", "E", "H"]
        widths = [16, 11, 18, 16, 18]
        for col, width in zip(cols, widths):
            mls.column_dimensions[col].width = width

        # add metadata
        meta = fmc_root['2: Matrix Least Squares Analysis']['metadata'].metadata
        mls.append([])  # Makes a new empty row
        for key, value in meta.items():
            mls.append([key, str(value)])
            cell = mls["A"+str(mls.max_row)]
            cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                       shrink_to_fit=False, indent=0)

        # add custom number formatting
        for i in range(5, mls.max_row):
            mls["D" + str(i)].number_format = "0.000 000 000"
            mls["H" + str(i)].number_format = "0.000 000 000"

    def add_weighing_dataset(self, se, cw_file,  nom, incl_datasets, cfg):
        """Adds relevant from each circular weighing for a given scheme entry

            Parameters
            ----------
            se : str
            cw_file : path
            nom : str
            incl_datasets : set
            cfg : configuration instance
        """
        if not os.path.isfile(cw_file):
            log.warning(f'No data yet collected for {se}')
        else:
            if len(se) > 30:
                # MS Excel has a limit of 31 characters in tab name
                sheet = self.wb.create_sheet(se[:27]+"...")
            else:
                sheet = self.wb.create_sheet(se)
            wt_grps = se.split()

            root = read(cw_file)
            log.info(f"Adding data for {se} from {cw_file}")
            sheet.append([f'Circular weighings for {se}'])
            sheet['A1'].font = Font(bold=True)
            sheet.append([f'Source file: {cw_file}'])

            try:
                root['Circular Weighings'][se]
            except KeyError:
                log.warning(f'No data yet collected for {se}')
                return

            first_dataset = True
            for dataset in root['Circular Weighings'][se].datasets():
                dname = dataset.name.split('_')

                if dname[0][-8:] == 'analysis':
                    run_id = 'run_' + dname[2]
                    run_no = float(run_id.strip('run_'))

                    weighdata = root.require_dataset(
                        root['Circular Weighings'][se].name + '/measurement_' + run_id)

                    try:
                        temps = weighdata.metadata.get("T" + IN_DEGREES_C).split(" to ")
                        rhs = weighdata.metadata.get("RH (%)").split(" to ")
                    except AttributeError:
                        try:
                            temps = weighdata.metadata["T range" + IN_DEGREES_C].split(" to ")
                            rhs = weighdata.metadata["RH (%)"].split(" to ")
                        except KeyError:
                            temps = []
                            rhs = []

                    if (str(float(nom)), se, dname[2]) in incl_datasets:
                        incl = 1
                        for t in temps:
                            self.collate_ambient['T' + IN_DEGREES_C].append(float(t))
                        for rh in rhs:
                            self.collate_ambient['RH (%)'].append(float(rh))
                    else:
                        incl = 0

                    # Get balance model number from balance alias:
                    bal_alias = weighdata.metadata.get("Balance")
                    bal_model = cfg.equipment[bal_alias].model
                    analysisdata = root.require_dataset(
                        root['Circular Weighings'][se].name + '/analysis_' + run_id)

                    if first_dataset:
                        # get information for header lines
                        sheet.append([])
                        sheet.append(['Balance', bal_model, 'Unit', weighdata.metadata.get("Unit")])
                        sheet.append([])
                        sheet.append(["Weight group loading order"])
                        pos = []
                        for key, value in weighdata.metadata.get("Weight group loading order").items():
                            sheet.append([key, value])
                            pos.append(key.strip("Position "))

                        header_pos = ["", "", ""]
                        for p in range(len(wt_grps) - 1):
                            key = pos[p] + ' - ' + pos[p+1]
                            header_pos.append(key)
                        header_pos.append(pos[-1] + ' - ' + pos[0])
                        sheet.append(header_pos)

                        header = [
                            "Timestamp",
                            'Run #',
                            "Included?"
                        ]
                        for row in analysisdata:  # get pairwise comparison names
                            header.append(f"({row[0]}) - ({row[1]})")
                        header += [
                            'Drift type',
                            'Residual Std. Dev.',
                            'Ambient OK?',
                            "min T" + IN_DEGREES_C,
                            "max T" + IN_DEGREES_C,
                            "min RH (%)",
                            "max RH (%)",
                            "start P (hPa)",
                            "end P (hPa)",
                            "Balance",
                            "Unit",
                        ]
                        sheet.append(header)

                        all_temps = []
                        all_rh = []

                        first_dataset = False

                    # add data for each run
                    data_list = [
                        weighdata.metadata.get("Mmt Timestamp"),
                        run_no,
                        incl
                    ]
                    for row in analysisdata:  # get pairwise comparison differences
                        data_list.append(row[2])

                    drift = analysisdata.metadata.get("Selected drift")
                    res_dict = eval(analysisdata.metadata.get("Residual std devs").replace("'", '"'))
                    res = res_dict[drift]

                    data_list += [
                        drift,
                        res,
                        str(weighdata.metadata.get("Ambient OK?"))
                    ]

                    if temps:
                        data_list += [
                            min(temps),
                            max(temps),
                            min(rhs),
                            max(rhs),
                        ]
                    else:
                        data_list += ["", "", "", ""]

                    try:
                        p = weighdata.metadata.get("Pressure (hPa)").split(" to ")
                        data_list += [p[0], p[1]]
                    except AttributeError:
                        data_list += ["", ""]
                    except TypeError:
                        data_list += ["", ""]

                    data_list += [bal_model, weighdata.metadata.get("Unit")]

                    sheet.append(data_list)

                    all_temps += temps
                    all_rh += rhs

            sheet['A6'].font = Font(italic=True)
            sheet.column_dimensions["A"].width = 15

            for cell in sheet[8+len(wt_grps)]:
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True,
                                        shrink_to_fit=False, indent=0)

            # determine formulae for mean and std dev of included weighings
            ave_row = ["", "", "Mean"]
            stdev_row = ["", "", "Std. Dev."]
            for col in string.ascii_lowercase[3:3+len(wt_grps)]:  # columns of weighing differences
                formula_ave = "=AVERAGE("
                formula_stdev = "=STDEV("
                for i, cell in enumerate(sheet['C']):
                    if cell.value == 1:  # included weighing
                        cell_name = col + str(i+1) + ','  # enumerate starts from 0, rows from 1
                        formula_ave += cell_name
                        formula_stdev += cell_name
                if formula_ave == "=AVERAGE(":
                    log.warning(f"No data selected for {se}")
                    break
                formula_ave = formula_ave.strip(',') + ")"
                formula_stdev = formula_stdev.strip(',') + ")"
                ave_row.append(formula_ave)
                stdev_row.append(formula_stdev)

            sheet.append([])
            sheet.append(ave_row)
            for cell in sheet[sheet.max_row]:
                cell.font = Font(italic=True)
            sheet.append(stdev_row)
            for cell in sheet[sheet.max_row]:
                cell.font = Font(italic=True)

            sheet.append([])
            if not all_temps:
                message = '<html>Please enter any known temperature values during weighing<br>' \
                          'for {}, separated by a space, ending {}</html>'.format(se, weighdata.metadata.get("Mmt Timestamp"))
                pt.show('text', message, font=FONTSIZE, title='Ambient Monitoring')
                reply = pt.wait_for_prompt_reply()
                try:
                    temperatures = reply.split()
                except AttributeError:
                    temperatures = []
                sheet.append(["Temperatures:"] + temperatures)
            if not all_rh:
                message = '<html>Please enter any known humidity values during weighing<br>' \
                          'for {}, separated by a space, ending {}</html>'.format(se, weighdata.metadata.get("Mmt Timestamp"))
                pt.show('text', message, font=FONTSIZE, title='Ambient Monitoring')
                reply = pt.wait_for_prompt_reply()
                try:
                    humidities = reply.split()
                except AttributeError:
                    humidities = []
                sheet.append(["Humidities:"] + humidities)

    def add_all_cwdata(self, cfg, incl_datasets,):
        scheme = self.wb['Scheme']
        i = self.first_scheme_entry_row  # Header rows occur before scheme

        while True:
            se = scheme.cell(row=i, column=1).value
            nom = scheme.cell(row=i, column=2).value
            if nom is None:
                break
            try:
                cw_file = os.path.join(cfg.folder,  f'{cfg.client}_{nom}.json')
            except TypeError:
                break
            self.add_weighing_dataset(se, cw_file, nom, incl_datasets, cfg)
            i += 1

    def add_overall_ambient(self):
        mls = self.wb["MLS Output Data"]
        mls.append([])
        mls.append(["Overall range of ambient conditions for included datasets (min, max)"])
        mls['A'+str(mls.max_row)].font = Font(italic=True)
        for key, value in self.collate_ambient.items():
            try:
                mls.append([key, min(value), max(value)])
            except ValueError:
                mls.append([key, "(no data)"])

    def save_workbook(self, folder, client):
        xl_output_file = os.path.join(folder, client + '_Summary.xlsx')
        # make backup
        if os.path.isfile(xl_output_file):
            back_up_folder = os.path.join(folder, "backups")
            new_index = len(os.listdir(back_up_folder))  # counts number of files in backup folder
            back_up_file = os.path.join(back_up_folder, client + '_summary_backup{}.xlsx'.format(new_index))
            os.rename(xl_output_file, back_up_file)  # this moves the file and renames it
        # protect each sheet
        for sheet in self.wb.sheetnames:
            self.wb[sheet].protection.password = 'Mass'
        self.wb.active = self.wb['Admin']
        # save the new file
        self.wb.save(xl_output_file)
        log.info('Data saved to {}'.format(xl_output_file))

        return xl_output_file
